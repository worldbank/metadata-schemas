import builtins
import os
import typing
import warnings
from typing import Any, Dict, List, Optional, Tuple, Type, Union, get_args, get_origin

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection
from pydantic import BaseModel

from .utils import (
    annotation_contains_list,
    get_subtype_of_optional_or_list,
    is_list_annotation,
    is_optional_annotation,
    seperate_simple_from_pydantic,
    subset_pydantic_model,
)

MAXCOL = 200


def protect_and_shade_given_cell(sheet, row: int, col: int):
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    sheet.cell(row=row, column=col).fill = grey_fill
    sheet.cell(row=row, column=col).protection = Protection(locked=True)


def protect_and_shade_row(sheet, row: int, colmin: int = 1, colmax: Optional[int] = None):
    if colmax is None:
        colmax = max(colmin, MAXCOL, sheet.max_column)
    for col in range(colmin, colmax):
        protect_and_shade_given_cell(sheet, row, col)


def protect_and_shade_col(sheet, col: int, rowmin: int, rowmax: int):
    for row in range(rowmin, rowmax):
        protect_and_shade_given_cell(sheet, row, col)


def unprotect_cell(sheet, row, column):
    sheet.cell(row=row, column=column).protection = Protection(locked=False)


def unprotect_row(sheet, row, colmin: int, colmax: Optional[int] = None):
    if colmax is None:
        colmax = max(colmin, MAXCOL, sheet.max_column)
    for col in range(colmin, colmax):
        unprotect_cell(sheet, row, col)


def unprotect_given_col(sheet, col: int, rowmin: int, rowmax: int):
    for row in range(rowmin, rowmax):
        unprotect_cell(sheet, row, col)


def shade_30_rows(doc_filepath: str, sheet_name: str, startrow: int):
    """For use after all data is written so there is a clear border around the data"""
    wb = load_workbook(doc_filepath)
    ws = wb[sheet_name]
    for r in range(startrow, startrow + 30):
        protect_and_shade_row(ws, r)
    wb.save(doc_filepath)


def correct_column_widths(filename: str, sheet_name: str):
    """
    Adjusts the column widths of an Excel sheet based on the maximum length of the content in each column.
    If a column has no filled values, its width remains unchanged.

    Args:
        filename (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to adjust column widths for.
    """
    # Load the existing workbook
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    # Adjust column widths based on the maximum length of the content in each column
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        if max_length > 0:  # Only adjust if there are filled values in the column
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(filename)


def shade_locked_cells(filename: str, sheet_name: str):
    """
    Shades every cell grey if it is locked and leaves it unshaded if it is not locked.

    Args:
        filename (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to apply the shading.
    """
    # Load the existing workbook
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    # Define the grey fill
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Iterate through each cell in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            if cell.protection.locked:
                cell.fill = grey_fill
            else:
                cell.fill = PatternFill()  # Remove any fill (reset to default)

    # Save the workbook
    wb.save(filename)


def write_to_cell(filename: str, sheet_name: str, row: int, col: int, text: str, isBold=False):
    """
    Writes text to a specified cell in the Excel file.

    Args:
        filename (str): The path to the Excel file.
        sheet_name (str): The name of the sheet.
        row_num (int): The row number (1-based index).
        col_num (int): The column number (1-based index).
        text (str): The text to write to the cell.
    """
    # Load the existing workbook or create a new one if it doesn't exist
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()

    # Select the worksheet by name
    ws = wb[sheet_name]

    # Write text to the specified cell
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=isBold)

    protect_and_shade_row(ws, row=row, colmin=col)

    # Save the workbook
    wb.save(filename)

    return row + 1


def create_sheet_and_write_title(doc_filepath: str, sheet_name: str, sheet_title: str, sheet_number: int = 0):
    """
    In the given excel document, creates a new sheet called sheet_name and in the top left cell
    writes in sheet_title in bold.

    It will create the excel document at doc_filepath if it does not already exist.

    The new sheet will be inserted at the specified sheet_number position. If sheet_number is
    greater than the total number of sheets, the new sheet will be added at the end.

    Args:
        doc_filepath (str): The path to the Excel document.
        sheet_name (str): The name of the new sheet to create.
        sheet_title (str): The title to write in the top left cell of the new sheet, in bold.
        sheet_number (int): The position to insert the new sheet (0-indexed). If greater than the
                            total number of sheets, the new sheet will be added at the end.

    Returns:
        int: index of next row below the final written row

    Raises:
        ValueError: if a sheet called sheet_name already exists in the document to prevent overwriting.
    """
    # Check if the file exists
    if os.path.exists(doc_filepath):
        workbook = load_workbook(doc_filepath)
    else:
        workbook = Workbook()
        # Remove the default sheet created by Workbook()
        if len(workbook.sheetnames) == 1 and workbook.sheetnames[0] == "Sheet":
            del workbook["Sheet"]

    # Check if the sheet already exists
    if sheet_name in workbook.sheetnames:
        raise ValueError(f"A sheet called '{sheet_name}' already exists in the document.")

    # Create a new sheet
    new_sheet = workbook.create_sheet(title=sheet_name)

    # Write the title in bold in the top left cell (A1)
    bold_font = Font(bold=True, size=14)
    new_sheet["A1"] = sheet_title
    new_sheet["A1"].font = bold_font

    # Shade the background of the cells in the first 2 rows grey and lock them
    for row in range(1, 3):
        protect_and_shade_row(new_sheet, row)

    # Determine the position to insert the new sheet
    total_sheets = len(workbook.sheetnames)
    insert_position = min(sheet_number, total_sheets)

    # Move the new sheet to the specified position
    workbook._sheets.insert(insert_position, workbook._sheets.pop())

    # Save the workbook
    workbook.save(doc_filepath)

    return 3


def replace_row_with_multiple_rows(original_df, new_df, row_to_replace):
    """
    Replace a specified row in the original DataFrame with multiple rows from the new DataFrame.

    Parameters:
    original_df (pd.DataFrame): The original DataFrame.
    new_df (pd.DataFrame): The new DataFrame with the rows to insert.
    row_to_replace (str): The name of the row to be replaced.

    Returns:
    pd.DataFrame: The updated DataFrame with the specified row replaced by the new rows.
    """
    # Split the original DataFrame into two parts: before and after the row to be replaced
    df_before = original_df.loc[:row_to_replace].iloc[:-1]
    df_after = original_df.loc[row_to_replace:].drop(row_to_replace, axis=0)

    # Concatenate the parts with the new rows
    df_replaced = pd.concat([df_before, new_df, df_after])
    df_replaced = df_replaced.dropna(how="all", axis=1)
    return df_replaced


def pydantic_to_dataframe(
    ob: Union[BaseModel, Dict, List[Dict]], annotations: Optional[Dict[str, typing._UnionGenericAlias]] = None
) -> Tuple[pd.DataFrame, List[int]]:
    """
    Convert to a dataframe, identifying rows that are made of lists and exploding them over multiple rows with
    hierarchical indices if needed.

    Returns the dataframe and also a list of the indexs (denoted by zero-based numbers) that are of list types.
    The list of indexs is intended to be used for appropriately shading the excel sheet.
    """
    if isinstance(ob, BaseModel):
        ob_dict = ob.model_dump(mode="json")
    else:
        ob_dict = ob
    df = pd.json_normalize(ob_dict).T
    print("pydantic_to_dataframe")
    print(df)
    list_indices = []
    if isinstance(ob, list):
        list_indices = list(range(len(df)))
    else:
        i = 0
        for idx in df.index:
            vals = df.loc[idx][0]
            field = ob_dict[idx.split(".")[0]]
            if (
                isinstance(vals, list)
                or (annotations is not None and annotation_contains_list(annotations[idx.split(".")[0]]))
                or (hasattr(field, "annotation") and annotation_contains_list(field.annotation))
            ):  # (hasattr(ob, "annotation") and annotation_contains_list(ob.annotation)):
                if vals is not None and len(vals) > 0 and (isinstance(vals[0], BaseModel) or isinstance(vals[0], Dict)):
                    print("list of base models", vals[0])
                    sub = pd.json_normalize(df.loc[idx].values[0]).reset_index(drop=True).T
                    sub.index = sub.index.map(lambda x: f"{idx}." + x)
                    df = replace_row_with_multiple_rows(df, sub, idx)
                    list_indices += list(range(i, i + len(sub)))
                    i += len(sub)
                else:
                    print("list of builtins or else empty")
                    df = replace_row_with_multiple_rows(
                        df, df.loc[idx].explode().to_frame().reset_index(drop=True).T, idx
                    )
                    list_indices.append(i)
                    i += 1
            else:
                i += 1
    print(df)
    if len(df):
        df.index = df.index.str.split(".", expand=True)
    return df, list_indices


def write_simple_pydantic_to_sheet(
    doc_filepath: str,
    sheet_name: str,
    ob: BaseModel,
    startrow: int,
    index_above=False,
    write_title=True,
    title: Optional[str] = None,
    annotations=None,
):
    """
    Assumes a pydantic object made up of built in types or pydantic objects utimately made of built in types or Lists.
    Do not use if the object or it's children contain  Dicts or enums.

    Starting from startrow, it writes the name of the pydantic object in the first column. It then writes the data
    starting in the row below and from the second column.

    If index_above = False then the data is printed with indexs down the second column and values down the third column
    If index_above = True then the data is printed with indexs along the second row and values along the third row

    Example:

        class Simple(BaseModel):
            a: str
            b: str

        example = Simple(a="value_a", b="value_b")
        # with index_above=True
        write_simple_pydantic_to_sheet("filename", "sheetname", example, startrow=1, index_above=True)

        gives:

            Simple
                   a            b
                   value_a      value_b

        # with index_above=False
        write_simple_pydantic_to_sheet("filename", "sheetname", example, startrow=1, index_above=False)

        gives:

            Simple

                    a     value_a
                    b     value_b

    Args:
        doc_filepath (str): The path to the Excel document.
        sheet_name (str): The name of the new sheet to create.
        ob (BaseModel): a pydantic class
        startrow (int): the row from which to start writing the data
        index_above (bool): if True then the index is written along a row with the data below, if False then the data is
            written in a column with the data to the right. Default is False

    Returns:
        int: index of next row below the final written row
    """
    if write_title:
        if title is None:
            title = ob.model_json_schema()["title"]
        startrow = write_to_cell(doc_filepath, sheet_name, startrow, 1, title, isBold=True)
    startcol = 2

    df, list_rows = pydantic_to_dataframe(ob=ob, annotations=annotations)
    index_levels = df.index.nlevels
    if index_above and index_levels > 1:
        warnings.warn(
            "Setting index_above=True is incompatible with a hierarchical index. Setting index_above to False.",
            UserWarning,
        )
        index_above = False

    if index_above:
        df = df.T

    # Annoyingly, openpyxl uses 1 based indexing but
    # But pandas uses 0 based indexing.

    with pd.ExcelWriter(doc_filepath, mode="a", if_sheet_exists="overlay") as writer:
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            header=index_above,
            index=not index_above,
            startrow=startrow - 1,
            startcol=startcol - 1,
            merge_cells=True,
        )

    # Open the Excel file with openpyxl
    workbook = load_workbook(doc_filepath)
    sheet = workbook[sheet_name]

    # Get the DataFrame dimensions
    rows, cols = df.shape

    if index_above:
        protect_and_shade_row(sheet, startrow)
        for c in range(startcol, cols + startcol):
            cell = sheet.cell(startrow, c)
            cell.font = Font(bold=False)
        for r in range(startrow + 1, startrow + rows + 1):
            unprotect_row(sheet, r, startcol, colmax=startcol + cols)
            protect_and_shade_row(sheet, r, colmin=startcol + cols)
        next_row = startrow + rows + 2
    else:
        for col in range(startcol, startcol + index_levels):
            protect_and_shade_col(sheet, col, startrow, startrow + rows)
            for r in range(startrow, startrow + rows):
                cell = sheet.cell(r, col)
                cell.font = Font(bold=False)
        firstdatacol = startcol + index_levels
        for i, r in enumerate(range(startrow, startrow + rows)):
            if i in list_rows:
                unprotect_row(sheet, r, firstdatacol)
            else:
                unprotect_row(sheet, r, firstdatacol, colmax=firstdatacol + 1)
                protect_and_shade_row(sheet, r, colmin=firstdatacol + 1)
        next_row = startrow + rows

    sheet.protection.enable()
    # Save the workbook
    workbook.save(doc_filepath)

    return next_row


def write_nested_simple_pydantic_to_sheet(
    doc_filepath: str, sheet_name: str, ob: BaseModel, startrow: int, index_above=False, write_title=True
):
    """
    Assumes the pydantic object is made up only of other pydantic objects that are themselves made up only of built in types
    """
    print(ob)
    children = seperate_simple_from_pydantic(ob)
    print(children["simple"])
    if len(children["simple"]):
        child_object = subset_pydantic_model(ob, children["simple"])
        startrow = write_simple_pydantic_to_sheet(
            doc_filepath,
            sheet_name,
            child_object,
            startrow,
            index_above=False,
            write_title=False,
            annotations={k: v.annotation for k, v in child_object.model_fields.items()},
        )
        print("Done with simple children, now nesting pydantic objects")
    if write_title:
        startrow = startrow + 2
    for mfield in children["pydantic"]:
        field = ob.model_dump(mode="json")[mfield]
        print(f"write_nested_simple_pydantic_to_sheet::428, field={field}")
        startrow = write_simple_pydantic_to_sheet(
            doc_filepath, sheet_name, field, startrow, index_above=index_above, title=mfield, write_title=write_title
        )

    return startrow


def write_across_many_sheets(doc_filepath: str, ob: BaseModel, title: Optional[str] = None):
    children = seperate_simple_from_pydantic(ob)
    sheet_number = 0
    if len(children["simple"]):
        if title is not None:
            title += " Metadata"
        else:
            title = "Metadata"
        sheet_name = "metadata"
        current_row = create_sheet_and_write_title(doc_filepath, sheet_name, f"{title}", sheet_number=sheet_number)
        child_object = subset_pydantic_model(ob, children["simple"])
        current_row = write_simple_pydantic_to_sheet(
            doc_filepath,
            sheet_name,
            child_object,
            current_row + 1,
            index_above=False,
            write_title=False,
            annotations={k: v.annotation for k, v in child_object.model_fields.items()},
        )
        correct_column_widths(doc_filepath, sheet_name=sheet_name)
        shade_30_rows(doc_filepath, sheet_name, current_row + 1)
        shade_locked_cells(doc_filepath, sheet_name)
        sheet_number += 1

    for fieldname in children["pydantic"]:
        print(f"\n\n{fieldname}\n")
        current_row = create_sheet_and_write_title(doc_filepath, fieldname, fieldname, sheet_number=sheet_number)
        field = getattr(ob, fieldname)
        if not isinstance(field, BaseModel):
            field = subset_pydantic_model(ob, [fieldname], name=fieldname)
            write_title = False
        else:
            write_title = True

        current_row = write_nested_simple_pydantic_to_sheet(
            doc_filepath, fieldname, field, current_row + 1, write_title=write_title
        )
        correct_column_widths(doc_filepath, sheet_name=fieldname)
        shade_30_rows(doc_filepath, fieldname, current_row + 1)
        shade_locked_cells(doc_filepath, fieldname)
        sheet_number += 1