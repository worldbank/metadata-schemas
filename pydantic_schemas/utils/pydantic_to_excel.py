import copy
import json
import os
import warnings
from enum import Enum
from typing import List, Optional, Tuple, Union, get_args

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import AnyUrl, BaseModel

from .schema_base_model import SchemaBaseModel
from .utils import (
    annotation_contains_dict,
    annotation_contains_list,
    assert_dict_annotation_is_strings_or_any,
    get_subtype_of_optional_or_list,
    is_list_annotation,
    is_optional_annotation,
    is_union_annotation,
    seperate_simple_from_pydantic,
    subset_pydantic_model,
)

MAXCOL = 30


def unprotect_cell(sheet, row, column):
    sheet.cell(row=row, column=column).protection = Protection(locked=False)


def unprotect_row(sheet, row, colmin: int, colmax: Optional[int] = None):
    if colmax is None:
        colmax = max(colmin, MAXCOL, sheet.max_column)
    for col in range(colmin, colmax + 1):
        unprotect_cell(sheet, row, col)


def unprotect_given_col(sheet, col: int, rowmin: int, rowmax: int):
    for row in range(rowmin, rowmax):
        unprotect_cell(sheet, row, col)


def protect_and_shade_given_cell(sheet, row: int, col: int):
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    sheet.cell(row=row, column=col).fill = grey_fill
    sheet.cell(row=row, column=col).protection = Protection(locked=True)


def protect_and_shade_row(sheet, row: int, colmin: int = 1, colmax: Optional[int] = None):
    if colmax is None:
        colmax = max(colmin, MAXCOL, sheet.max_column)
    for col in range(colmin, colmax):
        protect_and_shade_given_cell(sheet, row, col)


def protect_and_shade_col(sheet, col: int, rowmin: int, rowmax: int):
    for row in range(rowmin, rowmax):
        protect_and_shade_given_cell(sheet, row, col)


def shade_locked_cells(worksheet: Worksheet):
    """
    Shades every cell grey if it is locked and leaves it unshaded if it is not locked.

    Args:
        worksheet (Worksheet): The openPyxl Worksheet from an Excel file.
        sheet_name (str): The name of the sheet to apply the shading.
    """
    # Define the grey fill
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Iterate through each cell in the worksheet
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.protection.locked:
                cell.fill = grey_fill
            else:
                cell.fill = PatternFill()  # Remove any fill (reset to default)


def correct_column_widths(worksheet: Worksheet):
    """
    Adjusts the column widths of an Excel sheet based on the maximum length of the content in each column.
    If a column has no filled values, its width remains unchanged.

    Args:
        workbook (Workbook): The openPyxl Workbook of an Excel file.
        sheet_name (str): The name of the sheet to adjust column widths for.
    """
    # Load the existing workbook
    # Adjust column widths based on the maximum length of the content in each column
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            if column != "A":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        if max_length > 0:  # Only adjust if there are filled values in the column
            max_length = max(min(max_length, 28), 11)
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width


def shade_80_rows_and_protect_sheet(worksheet: Worksheet, startrow: int):
    """For use after all data is written so there is a clear border around the data"""
    for r in range(startrow, startrow + 80):
        protect_and_shade_row(worksheet, r)
    worksheet.protection = SheetProtection(
        sheet=True,
        formatCells=False,
        formatColumns=False,
        formatRows=False,
        insertColumns=False,
        insertRows=True,
        insertHyperlinks=False,
        deleteColumns=False,
        deleteRows=True,
        selectLockedCells=False,
        selectUnlockedCells=False,
    )


def replace_row_with_multiple_rows(original_df, new_df, row_to_replace):
    """
    Replace a specified row in the original DataFrame with multiple rows from the new DataFrame.

    Parameters:
    original_df (pd.DataFrame): The original DataFrame.
    new_df (pd.DataFrame): The new DataFrame with the rows to insert.
    row_to_replace (str): The name of the row to be replaced.

    Returns:
    pd.DataFrame: The updated DataFrame with the specified row replaced by the new rows.
    """
    # Split the original DataFrame into two parts: before and after the row to be replaced
    df_before = original_df.loc[:row_to_replace].iloc[:-1]
    df_after = original_df.loc[row_to_replace:].drop(row_to_replace, axis=0)

    # Concatenate the parts with the new rows
    return pd.concat([df_before, new_df, df_after])


def count_lists(model_fields, idx: str):
    """
    idx is a string name of a nested field seperated by dots like
        "identification_info.citation.alternateTitle"
    """
    n_lists = 0
    for part in idx.split("."):
        try:
            anno = model_fields[part].annotation
        except KeyError as e:
            raise KeyError(f"bad model fields given {idx}, for {part} of {model_fields}") from e
        n_lists += annotation_contains_list(anno)
        if is_optional_annotation(anno) or is_list_annotation(anno):
            anno = get_subtype_of_optional_or_list(anno)
        if hasattr(anno, "model_fields"):
            model_fields = anno.model_fields
        else:
            break
    return n_lists, anno


def pydantic_to_dataframe(
    ob: Union[BaseModel, List[BaseModel]],
    debug: bool = False,
) -> Tuple[pd.DataFrame, List[int]]:
    """
    Convert to a dataframe, identifying rows that are made of lists and exploding them over multiple rows with
    hierarchical indices if needed.

    Returns the dataframe and also a list of the indexs (denoted by zero-based numbers) that are of list types.
    The list of indexs is intended to be used for appropriately shading the excel sheet.
    """
    if isinstance(ob, list):
        ob_dict = [elem.model_dump() for elem in ob]
        annotations = {k: v.annotation for k, v in ob[0].model_fields.items()}
        model_fields = ob[0].model_fields
        is_list_of_objects = True
    else:
        ob_dict = ob.model_dump()
        annotations = {k: v.annotation for k, v in ob.model_fields.items()}
        model_fields = ob.model_fields
        is_list_of_objects = False
    df = pd.json_normalize(ob_dict).T
    if debug:
        print("pydantic_to_dataframe")
        print(df)

    i = 0
    list_indices = []
    observed_dicts = set()
    enums = {}
    for idx in df.index:
        if idx.split(".")[0] in observed_dicts:
            continue
        if debug:
            print(f"pydantic_to_dataframe::202 idx = {idx}, df = {df}")
        vals = df.loc[idx]  # [0]
        number_of_lists, anno = count_lists(model_fields, idx)
        number_of_lists = number_of_lists + int(is_list_of_objects)
        if debug:
            print(f"vals: {vals}")
            print(f'idx.split(".")[0]: {idx.split(".")[0]}')
            print(f'annotations[idx.split(".")[0]]: {annotations[idx.split(".")[0]]}')
            print(f"number of lists = {number_of_lists}")
            print(f"anno = {anno}")

        if annotation_contains_dict(anno):
            if debug:
                print(f"annotation contains dict, {ob_dict[idx.split('.')[0]]}")
            fieldname = idx.split(".")[0]
            subdf = df[df.index.str.startswith(f"{fieldname}.")]
            field = {"".join(i.split(".")[1:]): v[0] for i, v in zip(subdf.index, subdf.values)}
            if debug:
                print(f"field: {field}")
            if is_union_annotation(anno) and (len(subdf) == 0 or (field is not None and not isinstance(field, dict))):
                args = [a for a in get_args(anno) if a is not type(None)]
                anno = [a for a in args if not annotation_contains_dict(a)][0]
                if debug:
                    print(f"falling back to {anno}")
            else:
                if debug:
                    print("Found a dictionary")
                if is_list_of_objects:
                    continue
                assert_dict_annotation_is_strings_or_any(anno)

                if field is None or len(field) == 0:
                    dict_df = pd.DataFrame(["", ""], index=["key", "value"])
                else:
                    dict_df = pd.DataFrame([field.keys(), field.values()], index=["key", "value"])
                if debug:
                    print(f"created a dict_df:\n{dict_df}")
                dict_df.index = dict_df.index.map(lambda x, fieldname=fieldname: f"{fieldname}.{x}")
                df = df[~df.index.str.startswith(f"{fieldname}.")]
                df = df[df.index != fieldname]
                df = pd.concat([df, dict_df])
                list_indices += list(range(i, i + 2))
                i += 2
                observed_dicts.add(fieldname)
                continue

        if number_of_lists >= 1:  #: or annotation_contains_dict(annotations[idx.split(".")[0]]):
            # if number_of_lists > 0:
            subtype = anno
            if debug:
                print("subtype = ", subtype)
                print("isinstance(subtype, BaseModel)", isinstance(subtype, type(BaseModel)))
                print("isinstance(subtype, dict)", annotation_contains_dict(subtype))  # isinstance(subtype, dict))
            if number_of_lists >= 2 or is_list_of_objects:  # is_list_of_objects:
                if debug:
                    print("list of lists")
                list_indices.append(i)
                i += 1

            elif isinstance(subtype, type(BaseModel)):  # isinstance(subtype, type(dict))
                if debug:
                    print("list of base models", vals, vals[0])
                    print("experiment:", vals[0])
                    print("experiment:", pd.DataFrame(vals[0]).T)
                if vals[0] is None or isinstance(vals[0], list) and len(vals[0]) == 0:
                    vals[0] = [None]
                sub = pd.json_normalize(vals[0]).T
                if debug:
                    print(sub)
                sub.index = sub.index.map(lambda x, idx=idx: f"{idx}." + x)
                if debug:
                    print(sub)
                df = replace_row_with_multiple_rows(df, sub, idx)
                if debug:
                    print(df)
                if debug:
                    print(list_indices)
                list_indices += list(range(i, i + len(sub)))
                if debug:
                    print(list_indices)
                i += len(sub)
                if debug:
                    print("done with basemodel subtype")
            else:
                if debug:
                    print("list of builtins or else empty")
                if vals[0] is None or isinstance(vals[0], list) and len(vals[0]) == 0:
                    vals[0] = [None]
                sub = pd.DataFrame(vals[0]).T
                if len(sub.index) == 1:
                    sub.index = [idx]
                else:
                    sub.index = sub.index.map(lambda x, idx=idx: f"{idx}." + x)
                df = replace_row_with_multiple_rows(df, sub, idx)
                if debug:
                    print("new df:", df)
                list_indices.append(i)
                i += 1
        else:
            if isinstance(annotations[idx.split(".")[0]], type(Enum)):
                dropdown_options = [e.value for e in annotations[idx.split(".")[0]]]
                dropdown = DataValidation(
                    type="list",
                    formula1=f'"{",".join(dropdown_options)}"',
                    showDropDown=False,
                    allow_blank=True,
                    showErrorMessage=True,
                )
                enums[i] = dropdown
            i += 1
    if debug:
        print(df)
    if len(df):
        df.index = df.index.str.split(".", expand=True)
    if is_list_of_objects:
        list_indices = list(range(len(df)))
    return df, list_indices, enums


def stringify_enum(elem):
    if isinstance(elem, Enum):
        return str(elem.value)
    raise TypeError(f"{elem} is not an enum")


def stringify_cell_element(elem):
    if isinstance(elem, list):
        return json.dumps(elem, default=stringify_enum)
    if isinstance(elem, Enum):
        return str(elem.value)
    if isinstance(elem, dict):
        return json.dumps(elem, default=stringify_enum)
    if isinstance(elem, AnyUrl):
        return elem.unicode_string()
    return elem


def write_pydantic_to_excel(ws, ob, row_number, debug=False):
    df, list_rows, enums = pydantic_to_dataframe(ob, debug=debug)
    list_rows_tracker = {}
    list_of_enums_tracker = {}
    for i, r in enumerate(dataframe_to_rows(df, index=True, header=False)):
        if debug:
            print(r)
        if all(x is None for x in r):
            continue
        string_r = [stringify_cell_element(val) for val in r]
        string_r = [""] + string_r
        if debug:
            print("about to append", string_r)
        ws.append(string_r)
        for col in range(2, df.index.nlevels + 2):
            cell = ws.cell(row=row_number, column=col)
            cell.font = Font(bold=True)
            cell.border = Border(
                top=Side(border_style=None),
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                bottom=Side(border_style=None),
            )
            if cell.value is not None and cell.value != "":
                if debug:
                    print("turning on some borders")
                border_copy = copy.copy(cell.border)
                border_copy.top = Side(border_style="thin")
                cell.border = border_copy
        min_unprotected_cell = df.index.nlevels + 2
        max_unprotected_cell = None if i - 1 in list_rows else min_unprotected_cell
        unprotect_row(ws, row_number, colmin=min_unprotected_cell, colmax=max_unprotected_cell)
        if i - 1 in enums:
            dropdown = enums[i - 1]
            ws.add_data_validation(dropdown)
            for j in range(
                min_unprotected_cell, ws.max_column if max_unprotected_cell is None else min_unprotected_cell + 1
            ):
                dropdown.add(ws.cell(row_number, j))
        if max_unprotected_cell is None:
            list_rows_tracker[row_number] = ws.max_column
            if i - 1 in enums:
                list_of_enums_tracker[row_number] = dropdown
        row_number += 1

    for col in range(2, df.index.nlevels + 2):
        cell = ws.cell(row=row_number, column=col)
        border_copy = copy.copy(cell.border)
        border_copy.top = Side(border_style="thin")
        cell.border = border_copy

    return row_number + 1, list_rows_tracker, list_of_enums_tracker


def write_title_and_version_info(
    ws: Worksheet, sheet_title: Optional[str], version: Optional[str], protect_title=True
) -> int:
    if sheet_title is None:
        return 1
    if sheet_title is not None:
        sheet_title = sheet_title.replace("_", " ")
    ws.append([sheet_title, None, version])

    if sheet_title is not None:
        bold_font = Font(bold=True, size=14)
        ws["A1"].font = bold_font
        if protect_title == False:
            unprotect_row(ws, 1, colmin=1, colmax=1)

    if version is not None:
        version_font = Font(name="Consolas", size=9)
        ws["C1"].font = version_font

    ws.append([])
    return 3


def write_pydantic_to_sheet(worksheet: Worksheet, ob: BaseModel, current_row: int, debug: bool = False) -> int:
    children = seperate_simple_from_pydantic(ob)
    if debug:
        print("Children:")
        print(children)
    list_rows = {}
    enum_list_rows = {}

    if len(children["simple"]):
        child_object = subset_pydantic_model(ob, children["simple"])
        current_row, sub_list_rows, sub_list_enums = write_pydantic_to_excel(
            ws=worksheet, ob=child_object, row_number=current_row
        )
        list_rows.update(sub_list_rows)
        enum_list_rows.update(sub_list_enums)

    for mfield in children["pydantic"]:
        worksheet.append([mfield])
        worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1
        child_object = getattr(ob, mfield)
        current_row, sub_list_rows, sub_list_enums = write_pydantic_to_excel(
            ws=worksheet, ob=child_object, row_number=current_row, debug=debug
        )
        list_rows.update(sub_list_rows)
        enum_list_rows.update(sub_list_enums)

    for row, col in list_rows.items():
        unprotect_row(worksheet, row, colmin=col, colmax=None)
        if row in enum_list_rows:
            dropdown = enum_list_rows[row]
            for j in range(col, worksheet.max_column):
                dropdown.add(worksheet.cell(row, j))
    return current_row


def open_or_create_workbook(doc_filepath):
    if os.path.exists(doc_filepath):
        workbook = load_workbook(doc_filepath)
    else:
        workbook = Workbook()
        # Remove the default sheet created by Workbook()
        if len(workbook.sheetnames) == 1 and workbook.sheetnames[0] == "Sheet":
            del workbook["Sheet"]
    return workbook


def create_sheet(workbook, sheetname, sheet_number):
    # Check if the sheet already exists
    if sheetname in workbook.sheetnames:
        raise ValueError(f"A sheet called '{sheetname}' already exists in the document.")

    # Create a new sheet
    new_sheet = workbook.create_sheet(title=sheetname)

    # Determine the position to insert the new sheet
    total_sheets = len(workbook.sheetnames)
    insert_position = min(sheet_number, total_sheets)

    # Move the new sheet to the specified position
    workbook._sheets.insert(insert_position, workbook._sheets.pop())
    return new_sheet


def write_to_single_sheet(doc_filepath: str, ob: BaseModel, title: Optional[str] = None, verbose=False):
    model_default_name = ob.model_json_schema()["title"]
    if title is None:
        title = model_default_name
    wb = open_or_create_workbook(doc_filepath)
    ws = create_sheet(wb, "metadata", sheet_number=0)
    version = create_version(ob)
    current_row = write_title_and_version_info(ws, title, version, protect_title=False)
    current_row = write_pydantic_to_sheet(ws, ob, current_row, debug=verbose)
    correct_column_widths(worksheet=ws)
    shade_80_rows_and_protect_sheet(worksheet=ws, startrow=current_row)
    shade_locked_cells(worksheet=ws)
    wb.save(doc_filepath)


def create_version(ob: SchemaBaseModel):
    """
    Create a version string from the metadata_type and metadata_type_version attributes of a SchemaBaseModel.
    Optionally include the template_uid and template_name attributes if they are not None.

    Args:
        ob (SchemaBaseModel): The SchemaBaseModel object to generate the version string for.

    Returns:
        str: The version string.

    Example output:
        'metadata_type: dataset, metadata_type_version: 1.0'
        'metadata_type: dataset, metadata_type_version: 1.0, template_uid: 1234, template_name: My Template'
    """
    metadata_type = None
    if hasattr(ob, "_metadata_type__") and ob._metadata_type__ is not None:
        if isinstance(ob._metadata_type__, str):
            metadata_type = ob._metadata_type__
        elif hasattr(ob._metadata_type__, "default"):
            metadata_type = ob._metadata_type__.default
    if metadata_type is None:
        warnings.warn(f"No metadata_type found in the {type(ob)} object.", stacklevel=2)

    metadata_type_version = None
    if hasattr(ob, "_metadata_type_version__") and ob._metadata_type_version__ is not None:
        if isinstance(ob._metadata_type_version__, str):
            metadata_type_version = ob._metadata_type_version__
        elif hasattr(ob._metadata_type_version__, "default"):
            metadata_type_version = ob._metadata_type_version__.default
    if metadata_type_version is None:
        warnings.warn(f"No metadata_type_version found in the {type(ob)} object.", stacklevel=2)
    output = f"metadata_type: {metadata_type}, metadata_type_version: {metadata_type_version}"

    template_name = None
    if hasattr(ob, "_template_name__") and ob._template_name__ is not None:
        if isinstance(ob._template_name__, str):
            template_name = ob._template_name__
        elif hasattr(ob._template_name__, "default"):
            template_name = ob._template_name__.default

    template_uid = None
    if hasattr(ob, "_template_uid__") and ob._template_uid__ is not None:
        if isinstance(ob._template_uid__, str):
            template_uid = ob._template_uid__
        elif hasattr(ob._template_uid__, "default"):
            template_uid = ob._template_uid__.default

    if template_name is not None and template_uid is not None:
        output += f", template_uid: {template_uid}, template_name: {template_name}"
    return output


def parse_version(version: str):
    """
    Parse a version string into a dictionary of key-value pairs.

    Args:
        version (str): The version string to parse.

    Returns:
        dict: A dictionary of key-value pairs extracted from the version string.

    Example input:
        'metadata_type: dataset, metadata_type_version: 1.0, template_uid: 1234, template_name: My Template'

    Example output:
        {
            'metadata_type': 'dataset',
            'metadata_type_version': '1.0',
            'template_uid': '1234',
            'template_name': 'My Template'
        }
    """
    version_dict = {}
    version_info = version.split(",")
    if len(version_info) > 4:
        template_name_info = ",".join(version_info[3:])
        version_info = version_info[:3]
        version_info.append(template_name_info)
    for item in version_info:
        key_values = item.strip().split(":")
        key = key_values[0]
        value = ":".join(key_values[1:])
        version_dict[key.strip()] = value.strip()
    return version_dict


def write_across_many_sheets(doc_filepath: str, ob: SchemaBaseModel, title: Optional[str] = None, verbose=False):
    wb = open_or_create_workbook(doc_filepath)
    ws = create_sheet(wb, "metadata", sheet_number=0)
    version = create_version(ob)
    current_row = write_title_and_version_info(ws, title, version, protect_title=False)

    children = seperate_simple_from_pydantic(ob)
    if verbose:
        print(f"children: {children}")
    sheet_number = 0

    if len(children["simple"]):
        child_object = subset_pydantic_model(ob, children["simple"])

        current_row = write_pydantic_to_sheet(ws, child_object, current_row, debug=verbose)
    correct_column_widths(worksheet=ws)
    shade_80_rows_and_protect_sheet(worksheet=ws, startrow=current_row)
    shade_locked_cells(worksheet=ws)
    sheet_number += 1

    for fieldname in children["pydantic"]:
        if verbose:
            print(f"\n\n{fieldname}\n")
        child_object = getattr(ob, fieldname)
        if verbose:
            print(child_object)
        ws = create_sheet(wb, fieldname, sheet_number=sheet_number)
        if not isinstance(child_object, BaseModel):
            child_object = subset_pydantic_model(ob, [fieldname], name=fieldname)
            sheet_title = None
        else:
            sheet_title = fieldname
        current_row = write_title_and_version_info(ws, sheet_title, None, protect_title=True)
        current_row = write_pydantic_to_sheet(ws, child_object, current_row, debug=verbose)
        correct_column_widths(worksheet=ws)
        shade_80_rows_and_protect_sheet(worksheet=ws, startrow=current_row)
        shade_locked_cells(worksheet=ws)
        sheet_number += 1
    wb.save(doc_filepath)
