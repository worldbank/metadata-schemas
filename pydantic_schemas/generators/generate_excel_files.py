import os

import openpyxl

from pydantic_schemas.metadata_manager import MetadataManager


def compare_excel_files(file1, file2):
    # Load the workbooks
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # Get all sheet names
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames

    # Check if both workbooks have the same sheets
    if sheets1 != sheets2:
        # print("Sheet names do not match")
        # print(f"File1 sheets: {sheets1}")
        # print(f"File2 sheets: {sheets2}")
        return False

    # Iterate through each sheet
    for sheet_name in sheets1:
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]

        # Iterate through each cell in the sheet
        for row in ws1.iter_rows():
            for cell in row:
                cell_address = cell.coordinate
                if sheet_name == "metadata" and cell_address == "C1":
                    continue  # Skip comparison for cell C1 in 'metadata' sheet which only contains the versioning number

                differences = []
                if ws1[cell_address].value != ws2[cell_address].value:
                    differences.append(f"Value: {ws1[cell_address].value} != {ws2[cell_address].value}")
                if (
                    ws1[cell_address].font.name != ws2[cell_address].font.name
                    or ws1[cell_address].font.size != ws2[cell_address].font.size
                    or ws1[cell_address].font.bold != ws2[cell_address].font.bold
                    or ws1[cell_address].font.italic != ws2[cell_address].font.italic
                ):
                    differences.append(f"Font: {ws1[cell_address].font} != {ws2[cell_address].font}")
                if (
                    ws1[cell_address].fill.start_color.index != ws2[cell_address].fill.start_color.index
                    or ws1[cell_address].fill.end_color.index != ws2[cell_address].fill.end_color.index
                ):
                    differences.append(f"Fill: {ws1[cell_address].fill} != {ws2[cell_address].fill}")
                if (
                    ws1[cell_address].border.left.style != ws2[cell_address].border.left.style
                    or ws1[cell_address].border.right.style != ws2[cell_address].border.right.style
                    or ws1[cell_address].border.top.style != ws2[cell_address].border.top.style
                    or ws1[cell_address].border.bottom.style != ws2[cell_address].border.bottom.style
                ):
                    differences.append(f"Border: {ws1[cell_address].border} != {ws2[cell_address].border}")
                if (
                    ws1[cell_address].alignment.horizontal != ws2[cell_address].alignment.horizontal
                    or ws1[cell_address].alignment.vertical != ws2[cell_address].alignment.vertical
                ):
                    differences.append(f"Alignment: {ws1[cell_address].alignment} != {ws2[cell_address].alignment}")

                if differences:
                    # print(f"Differences found at {sheet_name} {cell_address}:")
                    # for difference in differences:
                    #     print(f"  - {difference}")
                    return False

    return True


metadata_manager = MetadataManager()

for metadata_name in metadata_manager.metadata_type_names:
    filename = f"excel_sheets/{metadata_name.capitalize()}_metadata.xlsx"
    print(f"Writing {metadata_name} outline to {filename}")
    if os.path.exists(filename):
        filename2 = f"excel_sheets/{metadata_name.capitalize()}_metadata2.xlsx"
        metadata_manager.write_metadata_outline_to_excel(metadata_name_or_class=metadata_name, filename=filename2)
        are_identical = compare_excel_files(filename, filename2)
        if are_identical:
            print("they're the same")
            os.remove(filename2)
        else:
            print("updating")
            os.remove(filename)
            os.rename(filename2, filename)
    else:
        metadata_manager.write_metadata_outline_to_excel(metadata_name_or_class=metadata_name, filename=filename)
    print()
